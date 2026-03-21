"""Manufacturing Tree API routes — Products, BOM, Work Centers, Production Orders."""

import io
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.session import get_db
from app.core.security import get_current_user, require_factory
from app.models.user import User
from app.schemas.manufacturing import (
    ProductCreate, ProductUpdate, ProductResponse,
    WorkCenterCreate, WorkCenterUpdate, WorkCenterResponse,
    BOMCreate, BOMUpdate, BOMResponse, BOMOperationCreate, BOMOperationResponse,
    ProductionOrderCreate, ProductionOrderUpdate, ProductionOrderResponse,
    ProductionOrderLineCreate, ProductionOrderLineResponse,
    POProductivitySummary,
)
from app.services.manufacturing_service import (
    ProductService, WorkCenterService, BOMService, ProductionOrderService,
)

router = APIRouter(prefix="/manufacturing", tags=["manufacturing"])


# ─── Products ─────────────────────────────────────────────────────────────────


@router.post("/products", response_model=ProductResponse)
async def create_product(
    data: ProductCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    return await ProductService.create(db, fid, data.model_dump())


@router.get("/products", response_model=list[ProductResponse])
async def list_products(
    active_only: bool = True,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    products = await ProductService.list_all(db, fid, active_only)
    return [ProductResponse.model_validate(p) for p in products]


@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    product = await ProductService.get(db, product_id, fid)
    if not product:
        raise HTTPException(404, "Product not found")
    return product


@router.patch("/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    data: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    product = await ProductService.update(db, product_id, fid, data.model_dump(exclude_unset=True))
    if not product:
        raise HTTPException(404, "Product not found")
    return product


# ─── Product Excel Template & Import ──────────────────────────────────────────


@router.get("/products/template/download")
async def download_product_template(
    user: User = Depends(get_current_user),
):
    """Download Excel template for bulk product import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["code", "name", "description", "unit_of_measure", "product_family"]
    col_widths = [15, 30, 40, 18, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Example rows
    examples = [
        ("SKU-001", "Widget Alpha", "Standard widget 50mm", "pcs", "Widgets"),
        ("SKU-002", "Bolt M8x30", "Hex bolt M8 x 30mm", "pcs", "Fasteners"),
        ("MAT-001", "Steel Sheet 2mm", "Cold rolled steel 1200x2400", "kg", "Raw Materials"),
    ]
    example_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    example_font = Font(italic=True, color="6B7280")
    for row_idx, example in enumerate(examples, 2):
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = example_fill
            cell.font = example_font
            cell.border = thin_border

    # Instructions sheet
    ws_info = wb.create_sheet("Instructions")
    ws_info.column_dimensions["A"].width = 25
    ws_info.column_dimensions["B"].width = 60
    instructions = [
        ("Column", "Description"),
        ("code *", "Unique product code / SKU (required)"),
        ("name *", "Product name (required)"),
        ("description", "Optional description"),
        ("unit_of_measure", "pcs, kg, liters, meters (default: pcs)"),
        ("product_family", "Optional grouping / category"),
        ("", ""),
        ("NOTES", ""),
        ("", "• Delete the example rows before uploading"),
        ("", "• code must be unique — duplicates will be skipped"),
        ("", "• Valid UOM values: pcs, kg, liters, meters"),
    ]
    for row_idx, (a, b) in enumerate(instructions, 1):
        cell_a = ws_info.cell(row=row_idx, column=1, value=a)
        cell_b = ws_info.cell(row=row_idx, column=2, value=b)
        if row_idx == 1:
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_template.xlsx"},
    )


@router.post("/products/import")
async def import_products_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Bulk import products from Excel file."""
    fid = require_factory(user)

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an Excel file (.xlsx)")

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
    except Exception:
        raise HTTPException(400, "Could not read Excel file. Ensure it's a valid .xlsx")

    # Get existing product codes to skip duplicates
    existing = await ProductService.list_all(db, fid, active_only=False)
    existing_codes = {p.code.upper() for p in existing}

    created = 0
    skipped = 0
    errors = []
    valid_uoms = {"pcs", "kg", "liters", "meters"}

    for row_idx, row in enumerate(rows, 2):
        if not row or not row[0] or not row[1]:
            continue  # Skip empty rows

        code = str(row[0]).strip()
        name = str(row[1]).strip()

        if not code or not name:
            errors.append(f"Row {row_idx}: code and name are required")
            continue

        if code.upper() in existing_codes:
            skipped += 1
            continue

        description = str(row[2]).strip() if row[2] else None
        uom = str(row[3]).strip().lower() if row[3] else "pcs"
        if uom not in valid_uoms:
            uom = "pcs"
        product_family = str(row[4]).strip() if len(row) > 4 and row[4] else None

        try:
            await ProductService.create(db, fid, {
                "code": code,
                "name": name,
                "description": description,
                "unit_of_measure": uom,
                "product_family": product_family,
            })
            existing_codes.add(code.upper())
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: validation or processing error")

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# ─── Work Centers ─────────────────────────────────────────────────────────────


@router.post("/work-centers", response_model=WorkCenterResponse)
async def create_work_center(
    data: WorkCenterCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    return await WorkCenterService.create(db, fid, data.model_dump())


@router.get("/work-centers", response_model=list[WorkCenterResponse])
async def list_work_centers(
    line_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    wcs = await WorkCenterService.list_all(db, fid, line_id)
    return [WorkCenterResponse.model_validate(wc) for wc in wcs]


@router.patch("/work-centers/{wc_id}", response_model=WorkCenterResponse)
async def update_work_center(
    wc_id: int,
    data: WorkCenterUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    wc = await WorkCenterService.update(db, wc_id, fid, data.model_dump(exclude_unset=True))
    if not wc:
        raise HTTPException(404, "Work center not found")
    return wc


# ─── BOM ──────────────────────────────────────────────────────────────────────


@router.post("/bom", response_model=BOMResponse)
async def create_bom(
    data: BOMCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    return await BOMService.create(db, fid, data.model_dump())


@router.get("/bom", response_model=list[BOMResponse])
async def list_boms(
    product_id: int | None = None,
    line_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    boms = await BOMService.list_all(db, fid, product_id, line_id)
    return [BOMResponse.model_validate(b) for b in boms]


@router.get("/bom/{bom_id}", response_model=BOMResponse)
async def get_bom(
    bom_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    bom = await BOMService.get(db, bom_id, fid)
    if not bom:
        raise HTTPException(404, "BOM not found")
    return bom


@router.get("/bom/for-line/{line_id}", response_model=list[BOMResponse])
async def get_boms_for_line(
    line_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    boms = await BOMService.get_active_for_line(db, fid, line_id)
    return [BOMResponse.model_validate(b) for b in boms]


@router.patch("/bom/{bom_id}/approve", response_model=BOMResponse)
async def approve_bom(
    bom_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    bom = await BOMService.approve(db, bom_id, fid, user.id)
    if not bom:
        raise HTTPException(404, "BOM not found")
    return BOMResponse.model_validate(bom)


# ─── BOM Excel Template / Import ─────────────────────────────────────────────


@router.get("/bom/template/download")
async def download_bom_template(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Download Excel template for BOM import."""
    fid = require_factory(user)

    # Get products and lines to show in template
    products = await ProductService.list_all(db, fid, active_only=True)
    from app.models.factory import ProductionLine
    from sqlalchemy import select
    lines_result = await db.execute(
        select(ProductionLine).where(ProductionLine.factory_id == fid)
    )
    lines = lines_result.scalars().all()

    wb = Workbook()

    # Sheet 1: BOM Template
    ws = wb.active
    ws.title = "BOM"
    headers = ["product_code *", "line_name *", "material_code *", "material_name *",
               "quantity_per_unit *", "unit_of_measure", "is_critical", "ideal_cycle_time_sec"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Example rows
    examples = [
        ("PROD-001", "Assembly Line 1", "MAT-001", "Steel Sheet 2mm", 2.5, "kg", "no", 45),
        ("PROD-001", "Assembly Line 1", "MAT-002", "M8 Bolt", 12, "pcs", "no", ""),
        ("PROD-001", "Assembly Line 1", "MAT-003", "Seal Ring", 1, "pcs", "yes", ""),
        ("PROD-002", "Assembly Line 1", "MAT-001", "Steel Sheet 2mm", 1.8, "kg", "no", 30),
    ]
    light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = light_fill
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    # Sheet 2: Reference — Products
    ws_prod = wb.create_sheet("Products (Reference)")
    ws_prod.cell(row=1, column=1, value="product_code").font = Font(bold=True)
    ws_prod.cell(row=1, column=2, value="product_name").font = Font(bold=True)
    for i, p in enumerate(products, 2):
        ws_prod.cell(row=i, column=1, value=p.code)
        ws_prod.cell(row=i, column=2, value=p.name)
    ws_prod.column_dimensions["A"].width = 20
    ws_prod.column_dimensions["B"].width = 40

    # Sheet 3: Reference — Lines
    ws_lines = wb.create_sheet("Lines (Reference)")
    ws_lines.cell(row=1, column=1, value="line_name").font = Font(bold=True)
    for i, ln in enumerate(lines, 2):
        ws_lines.cell(row=i, column=1, value=ln.name)
    ws_lines.column_dimensions["A"].width = 30

    # Sheet 4: Instructions
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ("Column", "Description"),
        ("product_code *", "Product code from your catalog (must exist)"),
        ("line_name *", "Production line name (must exist)"),
        ("material_code *", "Unique code for the material/component"),
        ("material_name *", "Material description"),
        ("quantity_per_unit *", "How many units consumed per 1 finished product"),
        ("unit_of_measure", "pcs, kg, liters, meters (default: pcs)"),
        ("is_critical", "yes/no — mark critical materials for line clearance"),
        ("ideal_cycle_time_sec", "Cycle time in seconds (only needed once per product+line)"),
        ("", ""),
        ("NOTES", ""),
        ("", "• Delete the example rows before uploading"),
        ("", "• Group rows by product_code + line_name"),
        ("", "• Multiple materials per product: one row per material"),
        ("", "• ideal_cycle_time_sec only needs to be on the first row of each group"),
        ("", "• If a BOM already exists for product+line, new materials will be added"),
        ("", "• Check the Products and Lines reference sheets for valid values"),
    ]
    for row_idx, (a, b) in enumerate(instructions, 1):
        cell_a = ws_info.cell(row=row_idx, column=1, value=a)
        cell_b = ws_info.cell(row=row_idx, column=2, value=b)
        if row_idx == 1:
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)
    ws_info.column_dimensions["A"].width = 25
    ws_info.column_dimensions["B"].width = 65

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bom_template.xlsx"},
    )


@router.post("/bom/import")
async def import_bom_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Bulk import BOM from Excel file."""
    fid = require_factory(user)

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an Excel file (.xlsx)")

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        raise HTTPException(400, "Could not read Excel file")

    # Load products and lines for lookup
    products = await ProductService.list_all(db, fid, active_only=True)
    product_map = {p.code.upper(): p.id for p in products}

    from app.models.factory import ProductionLine
    from sqlalchemy import select
    lines_result = await db.execute(
        select(ProductionLine).where(ProductionLine.factory_id == fid)
    )
    lines = lines_result.scalars().all()
    line_map = {ln.name.upper(): ln.id for ln in lines}

    # Group rows by (product_code, line_name)
    from collections import defaultdict
    groups: dict[tuple[str, str], list] = defaultdict(list)
    cycle_times: dict[tuple[str, str], float] = {}
    errors = []

    for row_idx, row in enumerate(rows, 2):
        if not row or not row[0] or not row[2]:
            continue

        product_code = str(row[0]).strip()
        line_name = str(row[1]).strip() if row[1] else ""
        material_code = str(row[2]).strip()
        material_name = str(row[3]).strip() if row[3] else ""
        qty = row[4]
        uom = str(row[5]).strip() if row[5] else "pcs"
        is_critical = str(row[6]).strip().lower() in ("yes", "true", "1", "si") if row[6] else False
        ct = row[7]

        if product_code.upper() not in product_map:
            errors.append(f"Row {row_idx}: product '{product_code}' not found in catalog")
            continue
        if line_name.upper() not in line_map:
            errors.append(f"Row {row_idx}: line '{line_name}' not found")
            continue
        if not material_name:
            errors.append(f"Row {row_idx}: material_name is required")
            continue
        try:
            qty_float = float(qty)
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx}: invalid quantity '{qty}'")
            continue

        key = (product_code.upper(), line_name.upper())
        groups[key].append({
            "material_code": material_code,
            "material_name": material_name,
            "quantity_per_unit": qty_float,
            "unit_of_measure": uom,
            "is_critical": is_critical,
        })
        if ct and key not in cycle_times:
            try:
                cycle_times[key] = float(ct)
            except (TypeError, ValueError):
                pass

    # Create BOMs
    boms_created = 0
    components_added = 0

    for (prod_code, line_name), materials in groups.items():
        product_id = product_map[prod_code]
        line_id = line_map[line_name]
        ct_sec = cycle_times.get((prod_code, line_name), 60.0)

        # Check if BOM already exists for this product+line
        existing_boms = await BOMService.list_all(db, fid, product_id, line_id)
        if existing_boms:
            # Add materials to existing BOM (skip duplicates by material_code)
            bom = existing_boms[0]
            existing_mat_codes = {c.material_code.upper() for c in bom.components if c.material_code}
            seq = max((c.sequence for c in bom.components), default=0)
            from app.models.manufacturing import BOMComponent
            for mat in materials:
                if mat["material_code"].upper() in existing_mat_codes:
                    continue
                seq += 1
                comp = BOMComponent(
                    bom_id=bom.id,
                    sequence=seq,
                    material_code=mat["material_code"],
                    material_name=mat["material_name"],
                    quantity_per_unit=mat["quantity_per_unit"],
                    unit_of_measure=mat["unit_of_measure"],
                    is_critical=mat["is_critical"],
                )
                db.add(comp)
                components_added += 1
        else:
            # Create new BOM
            bom_data = {
                "product_id": product_id,
                "production_line_id": line_id,
                "ideal_cycle_time_sec": ct_sec,
                "components": [
                    {**mat, "sequence": i + 1}
                    for i, mat in enumerate(materials)
                ],
            }
            try:
                await BOMService.create(db, fid, bom_data)
                boms_created += 1
                components_added += len(materials)
            except Exception as e:
                errors.append(f"BOM for {prod_code}/{line_name}: processing error")

    await db.commit()

    return {
        "boms_created": boms_created,
        "components_added": components_added,
        "errors": errors,
        "total_rows": len(rows),
    }


# ─── Production Orders ───────────────────────────────────────────────────────


@router.post("/orders", response_model=ProductionOrderResponse)
async def create_production_order(
    data: ProductionOrderCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.create(db, fid, user.id, data.model_dump(exclude_unset=True))
    return po


@router.get("/orders", response_model=list[ProductionOrderResponse])
async def list_production_orders(
    status: str | None = None,
    line_id: int | None = None,
    product_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    orders = await ProductionOrderService.list_all(db, fid, status, line_id, product_id)
    return [ProductionOrderResponse.model_validate(o) for o in orders]


@router.get("/orders/{order_id}", response_model=ProductionOrderResponse)
async def get_production_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.get(db, order_id, fid)
    if not po:
        raise HTTPException(404, "Production order not found")
    return po


@router.patch("/orders/{order_id}", response_model=ProductionOrderResponse)
async def update_production_order(
    order_id: int,
    data: ProductionOrderUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.get(db, order_id, fid)
    if not po:
        raise HTTPException(404, "Production order not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        if v is not None:
            setattr(po, k, v)
    await db.commit()
    await db.refresh(po)
    return ProductionOrderResponse.model_validate(po)


@router.post("/orders/{order_id}/release", response_model=ProductionOrderResponse)
async def release_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.update_status(db, order_id, fid, "released")
    if not po:
        raise HTTPException(404, "Production order not found")
    return ProductionOrderResponse.model_validate(po)


@router.post("/orders/{order_id}/start", response_model=ProductionOrderResponse)
async def start_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.get(db, order_id, fid)
    if not po:
        raise HTTPException(404, "Production order not found")
    if po.qc_hold:
        raise HTTPException(409, detail={
            "message": "Production order is under QC hold",
            "hold_reason": po.qc_hold_reason,
        })
    po = await ProductionOrderService.update_status(db, order_id, fid, "in_progress")
    return ProductionOrderResponse.model_validate(po)


@router.post("/orders/{order_id}/close", response_model=ProductionOrderResponse)
async def close_order(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.update_status(db, order_id, fid, "completed", user.id)
    if not po:
        raise HTTPException(404, "Production order not found")
    return ProductionOrderResponse.model_validate(po)


@router.post("/orders/{order_id}/hold", response_model=ProductionOrderResponse)
async def place_hold(
    order_id: int,
    reason: str = "Manual QC hold",
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.place_qc_hold(db, order_id, fid, reason)
    if not po:
        raise HTTPException(404, "Production order not found")
    return ProductionOrderResponse.model_validate(po)


@router.post("/orders/{order_id}/release-hold", response_model=ProductionOrderResponse)
async def release_hold(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    po = await ProductionOrderService.release_qc_hold(db, order_id, fid)
    if not po:
        raise HTTPException(404, "Production order not found")
    return ProductionOrderResponse.model_validate(po)


@router.get("/orders/{order_id}/summary", response_model=POProductivitySummary)
async def get_order_summary(
    order_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    fid = require_factory(user)
    summary = await ProductionOrderService.get_productivity_summary(db, order_id, fid)
    if not summary:
        raise HTTPException(404, "Production order not found")
    return summary
