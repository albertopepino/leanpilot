import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=14, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F46E5")
sub_header_font = Font(bold=True, size=11, color="1F2937")
sub_header_fill = PatternFill("solid", fgColor="E0E7FF")
input_font = Font(color="0000FF")
formula_font = Font(color="000000")
money_fmt = '$#,##0'
money_fmt_neg = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
num_fmt = '#,##0'
thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_subheader(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

# SHEET 1: Executive Summary
ws1 = wb.active
ws1.title = "Executive Summary"
data = [
    ["LEANPILOT - BUSINESS PLAN", "", "", ""],
    [],
    ["Company Overview", "", "", ""],
    ["Company Name", "LeanPilot SRL", "", ""],
    ["Founded", "2025", "", ""],
    ["Headquarters", "Italy (EU)", "", ""],
    ["Industry", "SaaS - Manufacturing Technology", "", ""],
    ["Website", "https://leanpilot.io", "", ""],
    [],
    ["Mission", "", "", ""],
    ["", "Democratize lean manufacturing for SME factories worldwide through AI-powered digital tools.", "", ""],
    [],
    ["Vision", "", "", ""],
    ["", "Become the global standard for digital lean management in factories with 10-500 employees.", "", ""],
    [],
    ["Value Proposition", "", "", ""],
    ["", "17 integrated lean tools + AI copilot in one platform, replacing fragmented Excel/paper processes.", "", ""],
    [],
    ["Key Differentiators", "", "", ""],
    ["1.", "All-in-one: OEE, TPM, SMED, VSM, 6S, CILT, Kaizen, A3, 5Why, Ishikawa, Pareto, Gemba, Andon + more", "", ""],
    ["2.", "AI-powered Factory Copilot (lean-only, multilingual)", "", ""],
    ["3.", "Lean Assessment wizard with personalized tool recommendations", "", ""],
    ["4.", "Mobile-first, designed for shop floor use", "", ""],
    ["5.", "GDPR + ePrivacy compliant (EU-first approach)", "", ""],
    [],
    ["Target Market", "", "", ""],
    ["Primary", "SME manufacturers (10-500 employees) in EU, Balkans, UK", "", ""],
    ["Secondary", "Lean consultants and training organizations", "", ""],
    ["Tertiary", "Large enterprises seeking departmental lean tools", "", ""],
    [],
    ["Business Model", "", "", ""],
    ["Type", "B2B SaaS - Monthly/Annual subscription", "", ""],
    ["Free Trial", "14 days, full access, no credit card", "", ""],
    ["Pricing Tiers", "Starter / Professional / Business / Enterprise", "", ""],
    ["Revenue Streams", "Subscriptions, AI add-on, custom integrations, training", "", ""],
]
for i, row in enumerate(data, 1):
    for j, val in enumerate(row, 1):
        cell = ws1.cell(row=i, column=j, value=val)
        cell.border = thin_border
        if i == 1:
            cell.font = Font(bold=True, size=16, color="4F46E5")
        elif val in ["Company Overview", "Mission", "Vision", "Value Proposition", "Key Differentiators", "Target Market", "Business Model"]:
            cell.font = Font(bold=True, size=12, color="4F46E5")
            cell.fill = PatternFill("solid", fgColor="EEF2FF")
auto_width(ws1)
# SHEET 2: Market Analysis
ws2 = wb.create_sheet("Market Analysis")
ws2.cell(row=1, column=1, value="MARKET ANALYSIS").font = Font(bold=True, size=14, color="4F46E5")
headers = ["Metric", "Value", "Basis", "Source"]
for j, h in enumerate(headers, 1):
    ws2.cell(row=3, column=j, value=h)
style_header(ws2, 3, 4)
market_data = [
    ["Global Manufacturing SaaS TAM", 50000000000, "Global market size 2025", "Source: Grand View Research, 2024"],
    ["EU SME Manufacturing SAM", 2500000000, "EU SME segment (~5% of TAM)", "Source: Eurostat SME data, 2024"],
    ["Lean-specific SOM (Year 5)", None, "Target 0.5% of SAM", "Calculated"],
    ["Target Factories (EU+Balkans)", 180000, "Factories 10-500 employees", "Source: Eurostat SBS, 2024"],
    ["Addressable (digital-ready)", 54000, "~30% digital adoption rate", "Source: EU Digital Economy Index"],
]
for i, row in enumerate(market_data, 4):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.border = thin_border
        if j == 2 and isinstance(val, (int, float)) and val:
            cell.number_format = money_fmt
            cell.font = input_font
ws2.cell(row=6, column=2, value="=B5*0.005").font = formula_font
ws2.cell(row=6, column=2).number_format = money_fmt

ws2.cell(row=11, column=1, value="COMPETITOR LANDSCAPE").font = Font(bold=True, size=12, color="4F46E5")
comp_headers = ["Competitor", "Focus", "Pricing", "Weakness vs LeanPilot"]
for j, h in enumerate(comp_headers, 1):
    ws2.cell(row=12, column=j, value=h)
style_subheader(ws2, 12, 4)
competitors = [
    ["Tulip", "No-code manufacturing apps", "$250+/mo", "Complex setup, no lean methodology focus"],
    ["Poka", "Connected worker platform", "Enterprise only", "No lean tools, large enterprise focus"],
    ["Evocon", "OEE monitoring", "EUR99+/mo", "OEE-only, no problem-solving tools"],
    ["Parsable", "Digital work instructions", "Enterprise", "No lean assessment, no AI copilot"],
    ["MaintainX", "CMMS / maintenance", "$16+/user/mo", "Maintenance-only, no lean suite"],
    ["SafetyCulture", "Inspections & audits", "$24+/user/mo", "Audit-focused, not lean-specific"],
    ["Excel/Paper", "Manual lean tracking", "Free", "No automation, no analytics, error-prone"],
]
for i, row in enumerate(competitors, 13):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val).border = thin_border
ws2.cell(row=22, column=1, value="SWOT ANALYSIS").font = Font(bold=True, size=12, color="4F46E5")
swot_headers = ["Strengths", "Weaknesses", "Opportunities", "Threats"]
for j, h in enumerate(swot_headers, 1):
    ws2.cell(row=23, column=j, value=h)
style_subheader(ws2, 23, 4)
swot = [
    ["All-in-one lean suite (17 tools)", "New brand, no track record", "EU Green Deal driving lean adoption", "Large players adding lean modules"],
    ["AI-powered copilot", "Small initial team", "Post-COVID digital transformation", "Open-source lean tools emerging"],
    ["GDPR-first design", "Limited initial languages", "Balkans underserved market", "Economic downturn reducing IT spend"],
    ["Competitive pricing", "No mobile app (PWA only)", "Lean consulting partnerships", "Resistance to digital tools in SMEs"],
    ["Multilingual (EN/IT, expanding)", "Single cloud region initially", "Government digitization subsidies", "Data sovereignty concerns"],
]
for i, row in enumerate(swot, 24):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val).border = thin_border
auto_width(ws2)

# SHEET 3: Revenue Projections
ws3 = wb.create_sheet("Revenue Projections")
ws3.cell(row=1, column=1, value="5-YEAR REVENUE PROJECTIONS").font = Font(bold=True, size=14, color="4F46E5")
years = ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
for j, h in enumerate(years, 1):
    ws3.cell(row=3, column=j, value=h)
style_header(ws3, 3, 6)
# Revenue rows layout
rev_labels = [
    ("New Customers (Starter)", [40, 120, 280, 500, 750]),
    ("New Customers (Professional)", [15, 60, 180, 350, 550]),
    ("New Customers (Business)", [5, 25, 80, 180, 320]),
    ("New Customers (Enterprise)", [1, 5, 15, 40, 80]),
    ("Total New Customers", None),
    ("Cumulative Customers", None),
    ("Churn Rate (Annual)", [0.15, 0.12, 0.10, 0.08, 0.07]),
    ("Churned Customers", None),
    ("Active Customers (End)", None),
    (None, None),
    ("PRICING (Monthly)", None),
    ("Starter", [49, 49, 59, 59, 69]),
    ("Professional", [149, 149, 179, 179, 199]),
    ("Business", [349, 349, 399, 399, 449]),
    ("Enterprise", [899, 899, 999, 999, 1099]),
    (None, None),
    ("MONTHLY RECURRING REVENUE", None),
    ("MRR - Starter", None),
    ("MRR - Professional", None),
    ("MRR - Business", None),
    ("MRR - Enterprise", None),
    ("Total MRR", None),
    ("ARR (Annual Recurring Revenue)", None),
    (None, None),
    ("ADDITIONAL REVENUE", None),
    ("AI Copilot Add-on (20% uptake)", None),
    ("Training & Onboarding", [5000, 20000, 60000, 120000, 200000]),
    ("Custom Integrations", [0, 10000, 40000, 100000, 200000]),
    ("Total Additional Revenue", None),
    (None, None),
    ("TOTAL ANNUAL REVENUE", None),
]
r = 4
for item in rev_labels:
    label, values = item
    if label is None:
        r += 1
        continue
    ws3.cell(row=r, column=1, value=label).border = thin_border
    if label in ["PRICING (Monthly)", "MONTHLY RECURRING REVENUE", "ADDITIONAL REVENUE"]:
        ws3.cell(row=r, column=1).font = Font(bold=True, size=11, color="4F46E5")
        ws3.cell(row=r, column=1).fill = PatternFill("solid", fgColor="EEF2FF")
    if values:
        for j, v in enumerate(values, 2):
            cell = ws3.cell(row=r, column=j, value=v)
            cell.border = thin_border
            cell.font = input_font
            if isinstance(v, float) and v < 1:
                cell.number_format = pct_fmt
            elif isinstance(v, int) and v >= 49:
                cell.number_format = money_fmt
            else:
                cell.number_format = num_fmt
    r += 1

# Row references
R_STARTER = 4; R_PROF = 5; R_BUS = 6; R_ENT = 7
R_TOTAL_NEW = 8; R_CUM = 9; R_CHURN_RATE = 10; R_CHURNED = 11; R_ACTIVE = 12
R_P_STARTER = 15; R_P_PROF = 16; R_P_BUS = 17; R_P_ENT = 18
R_MRR_S = 21; R_MRR_P = 22; R_MRR_B = 23; R_MRR_E = 24
R_TOTAL_MRR = 25; R_ARR = 26
R_AI_ADDON = 29; R_TRAINING = 30; R_CUSTOM = 31; R_TOTAL_ADD = 32
R_TOTAL_REV = 34
for col in range(2, 7):
    c = get_column_letter(col)
    prev = get_column_letter(col - 1) if col > 2 else None
    # Total New Customers
    ws3.cell(row=R_TOTAL_NEW, column=col, value=f"=SUM({c}{R_STARTER}:{c}{R_ENT})").font = formula_font
    ws3.cell(row=R_TOTAL_NEW, column=col).number_format = num_fmt
    ws3.cell(row=R_TOTAL_NEW, column=col).border = thin_border
    # Cumulative
    if col == 2:
        ws3.cell(row=R_CUM, column=col, value=f"={c}{R_TOTAL_NEW}").font = formula_font
    else:
        ws3.cell(row=R_CUM, column=col, value=f"={prev}{R_CUM}+{c}{R_TOTAL_NEW}").font = formula_font
    ws3.cell(row=R_CUM, column=col).number_format = num_fmt
    ws3.cell(row=R_CUM, column=col).border = thin_border
    # Churned
    if col == 2:
        ws3.cell(row=R_CHURNED, column=col, value=f"=ROUND({c}{R_CUM}*{c}{R_CHURN_RATE},0)").font = formula_font
    else:
        ws3.cell(row=R_CHURNED, column=col, value=f"=ROUND({prev}{R_ACTIVE}*{c}{R_CHURN_RATE},0)").font = formula_font
    ws3.cell(row=R_CHURNED, column=col).number_format = num_fmt
    ws3.cell(row=R_CHURNED, column=col).border = thin_border
    # Active End
    ws3.cell(row=R_ACTIVE, column=col, value=f"={c}{R_CUM}-{c}{R_CHURNED}").font = formula_font
    ws3.cell(row=R_ACTIVE, column=col).number_format = num_fmt
    ws3.cell(row=R_ACTIVE, column=col).border = thin_border
    # MRR formulas
    ws3.cell(row=R_MRR_S, column=col, value=f"={c}{R_STARTER}*{c}{R_P_STARTER}").font = formula_font
    ws3.cell(row=R_MRR_S, column=col).number_format = money_fmt
    ws3.cell(row=R_MRR_S, column=col).border = thin_border
    ws3.cell(row=R_MRR_P, column=col, value=f"={c}{R_PROF}*{c}{R_P_PROF}").font = formula_font
    ws3.cell(row=R_MRR_P, column=col).number_format = money_fmt
    ws3.cell(row=R_MRR_P, column=col).border = thin_border
    ws3.cell(row=R_MRR_B, column=col, value=f"={c}{R_BUS}*{c}{R_P_BUS}").font = formula_font
    ws3.cell(row=R_MRR_B, column=col).number_format = money_fmt
    ws3.cell(row=R_MRR_B, column=col).border = thin_border
    ws3.cell(row=R_MRR_E, column=col, value=f"={c}{R_ENT}*{c}{R_P_ENT}").font = formula_font
    ws3.cell(row=R_MRR_E, column=col).number_format = money_fmt
    ws3.cell(row=R_MRR_E, column=col).border = thin_border
    # Total MRR
    ws3.cell(row=R_TOTAL_MRR, column=col, value=f"=SUM({c}{R_MRR_S}:{c}{R_MRR_E})").font = Font(bold=True)
    ws3.cell(row=R_TOTAL_MRR, column=col).number_format = money_fmt
    ws3.cell(row=R_TOTAL_MRR, column=col).border = thin_border
    # ARR
    ws3.cell(row=R_ARR, column=col, value=f"={c}{R_TOTAL_MRR}*12").font = Font(bold=True)
    ws3.cell(row=R_ARR, column=col).number_format = money_fmt
    ws3.cell(row=R_ARR, column=col).border = thin_border
    # AI Add-on
    ws3.cell(row=R_AI_ADDON, column=col, value=f"=ROUND({c}{R_ACTIVE}*0.2*29*12,0)").font = formula_font
    ws3.cell(row=R_AI_ADDON, column=col).number_format = money_fmt
    ws3.cell(row=R_AI_ADDON, column=col).border = thin_border
    # Total Additional
    ws3.cell(row=R_TOTAL_ADD, column=col, value=f"=SUM({c}{R_AI_ADDON}:{c}{R_CUSTOM})").font = formula_font
    ws3.cell(row=R_TOTAL_ADD, column=col).number_format = money_fmt
    ws3.cell(row=R_TOTAL_ADD, column=col).border = thin_border
    # TOTAL ANNUAL REVENUE
    ws3.cell(row=R_TOTAL_REV, column=col, value=f"={c}{R_ARR}+{c}{R_TOTAL_ADD}").font = Font(bold=True, size=12, color="4F46E5")
    ws3.cell(row=R_TOTAL_REV, column=col).number_format = money_fmt
    ws3.cell(row=R_TOTAL_REV, column=col).border = thin_border
    ws3.cell(row=R_TOTAL_REV, column=col).fill = PatternFill("solid", fgColor="EEF2FF")

ws3.cell(row=R_TOTAL_REV, column=1).font = Font(bold=True, size=12, color="4F46E5")
ws3.cell(row=R_TOTAL_REV, column=1).fill = PatternFill("solid", fgColor="EEF2FF")
auto_width(ws3)
# SHEET 4: Cost Structure
ws4 = wb.create_sheet("Cost Structure")
ws4.cell(row=1, column=1, value="5-YEAR COST STRUCTURE").font = Font(bold=True, size=14, color="4F46E5")
for j, h in enumerate(years, 1):
    ws4.cell(row=3, column=j, value=h)
style_header(ws4, 3, 6)

costs = [
    ("INFRASTRUCTURE", None),
    ("Cloud Hosting (AWS/Vercel)", [6000, 18000, 48000, 96000, 180000]),
    ("Database & Redis", [2400, 6000, 14400, 28800, 48000]),
    ("OpenAI API Costs", [3600, 12000, 36000, 84000, 156000]),
    ("Domain, SSL, CDN", [500, 500, 1000, 1000, 2000]),
    ("Monitoring & Security", [1200, 3600, 7200, 12000, 18000]),
    ("Subtotal Infrastructure", None),
    (None, None),
    ("TEAM (Salaries + Benefits)", None),
    ("Founders (2x)", [0, 60000, 120000, 180000, 240000]),
    ("Backend Engineers", [0, 45000, 135000, 225000, 360000]),
    ("Frontend Engineers", [0, 45000, 90000, 180000, 270000]),
    ("DevOps / SRE", [0, 0, 55000, 110000, 165000]),
    ("Product / Design", [0, 0, 50000, 100000, 150000]),
    ("Sales & Marketing", [0, 30000, 90000, 180000, 300000]),
    ("Customer Success", [0, 0, 45000, 90000, 180000]),
    ("Subtotal Team", None),
    (None, None),
    ("MARKETING & SALES", None),
    ("Content Marketing & SEO", [3000, 12000, 30000, 60000, 96000]),
    ("Paid Advertising", [2000, 15000, 50000, 100000, 180000]),
    ("Events & Trade Shows", [0, 5000, 20000, 40000, 60000]),
    ("Partner Commissions", [0, 3000, 15000, 40000, 80000]),
    ("Subtotal Marketing", None),
    (None, None),
    ("OPERATIONS", None),
    ("Legal & Compliance", [3000, 6000, 12000, 18000, 24000]),
    ("Accounting & Admin", [2000, 5000, 10000, 18000, 24000]),
    ("Insurance", [1500, 3000, 6000, 10000, 15000]),
    ("Office / Coworking", [0, 6000, 12000, 24000, 36000]),
    ("Subtotal Operations", None),
    (None, None),
    ("TOTAL COSTS", None),
    ("Revenue (from Projections)", None),
    ("NET PROFIT / (LOSS)", None),
    ("Profit Margin", None),
]

r = 4
subtotal_rows = {}
category_start = {}
current_cat = None
total_cost_row = None
revenue_row = None
profit_row = None

for item in costs:
    label, values = item
    if label is None:
        r += 1
        continue
    ws4.cell(row=r, column=1, value=label).border = thin_border
    if label in ["INFRASTRUCTURE", "TEAM (Salaries + Benefits)", "MARKETING & SALES", "OPERATIONS"]:
        ws4.cell(row=r, column=1).font = Font(bold=True, size=11, color="4F46E5")
        ws4.cell(row=r, column=1).fill = PatternFill("solid", fgColor="EEF2FF")
        current_cat = label
        category_start[current_cat] = r + 1
    elif label.startswith("Subtotal"):
        subtotal_rows[current_cat] = r
        for col in range(2, 7):
            cc = get_column_letter(col)
            start_r = category_start[current_cat]
            ws4.cell(row=r, column=col, value=f"=SUM({cc}{start_r}:{cc}{r-1})").font = Font(bold=True)
            ws4.cell(row=r, column=col).number_format = money_fmt
            ws4.cell(row=r, column=col).border = thin_border
    elif label == "TOTAL COSTS":
        ws4.cell(row=r, column=1).font = Font(bold=True, size=12, color="4F46E5")
        ws4.cell(row=r, column=1).fill = PatternFill("solid", fgColor="EEF2FF")
        for col in range(2, 7):
            cc = get_column_letter(col)
            refs = "+".join([f"{cc}{subtotal_rows[cat]}" for cat in subtotal_rows])
            ws4.cell(row=r, column=col, value=f"={refs}").font = Font(bold=True, size=12)
            ws4.cell(row=r, column=col).number_format = money_fmt
            ws4.cell(row=r, column=col).border = thin_border
            ws4.cell(row=r, column=col).fill = PatternFill("solid", fgColor="EEF2FF")
        total_cost_row = r
    elif label == "Revenue (from Projections)":
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws4.cell(row=r, column=col, value=f"='Revenue Projections'!{cc}{R_TOTAL_REV}").font = formula_font
            ws4.cell(row=r, column=col).number_format = money_fmt
            ws4.cell(row=r, column=col).border = thin_border
        revenue_row = r
    elif label == "NET PROFIT / (LOSS)":
        ws4.cell(row=r, column=1).font = Font(bold=True, size=12, color="4F46E5")
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws4.cell(row=r, column=col, value=f"={cc}{revenue_row}-{cc}{total_cost_row}").font = Font(bold=True, size=12)
            ws4.cell(row=r, column=col).number_format = money_fmt_neg
            ws4.cell(row=r, column=col).border = thin_border
        profit_row = r
    elif label == "Profit Margin":
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws4.cell(row=r, column=col, value=f"=IF({cc}{revenue_row}=0,0,{cc}{profit_row}/{cc}{revenue_row})").font = formula_font
            ws4.cell(row=r, column=col).number_format = pct_fmt
            ws4.cell(row=r, column=col).border = thin_border
    elif values:
        for j, v in enumerate(values, 2):
            cell = ws4.cell(row=r, column=j, value=v)
            cell.border = thin_border
            cell.font = input_font
            cell.number_format = money_fmt
    r += 1
auto_width(ws4)

# SHEET 5: Pricing Strategy
ws5 = wb.create_sheet("Pricing Strategy")
ws5.cell(row=1, column=1, value="PRICING STRATEGY & FEATURE MATRIX").font = Font(bold=True, size=14, color="4F46E5")
pricing_headers = ["Feature / Capability", "Starter\n$49/mo", "Professional\n$149/mo", "Business\n$349/mo", "Enterprise\n$899/mo"]
for j, h in enumerate(pricing_headers, 1):
    ws5.cell(row=3, column=j, value=h)
    ws5.cell(row=3, column=j).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
style_header(ws5, 3, 5)
ws5.row_dimensions[3].height = 40
features = [
    ("CORE LEAN TOOLS", None, None, None, None),
    ("OEE Dashboard", "Basic", "Advanced", "Advanced", "Custom"),
    ("Production Input", "Y", "Y", "Y", "Y"),
    ("Hourly Production Board", "Y", "Y", "Y", "Y"),
    ("Andon Board", "-", "Y", "Y", "Y"),
    ("6S Audit", "Y", "Y", "Y", "Y"),
    ("CILT Checklist", "-", "Y", "Y", "Y"),
    ("TPM Dashboard", "-", "Y", "Y", "Y"),
    (None, None, None, None, None),
    ("PROBLEM SOLVING", None, None, None, None),
    ("5 Why Analysis", "Y", "Y", "Y", "Y"),
    ("Ishikawa Diagram", "Y", "Y", "Y", "Y"),
    ("Pareto Chart", "-", "Y", "Y", "Y"),
    ("A3 Report", "-", "Y", "Y", "Y"),
    (None, None, None, None, None),
    ("IMPROVEMENT", None, None, None, None),
    ("Kaizen Board", "-", "Y", "Y", "Y"),
    ("Value Stream Mapping", "-", "-", "Y", "Y"),
    ("SMED Tracker", "-", "-", "Y", "Y"),
    ("Gemba Walk", "-", "Y", "Y", "Y"),
    (None, None, None, None, None),
    ("AI & ANALYTICS", None, None, None, None),
    ("Factory Copilot (AI)", "-", "Basic", "Advanced", "Unlimited"),
    ("Lean Assessment", "Y", "Y", "Y", "Y"),
    ("Custom Reports", "-", "-", "Y", "Y"),
    ("API Access", "-", "-", "-", "Y"),
    (None, None, None, None, None),
    ("SUPPORT & ADMIN", None, None, None, None),
    ("Users Included", "3", "10", "25", "Unlimited"),
    ("Production Lines", "1", "3", "10", "Unlimited"),
    ("Data Retention", "90 days", "1 year", "3 years", "Unlimited"),
    ("Support", "Email", "Priority", "Dedicated", "24/7 + SLA"),
    ("Onboarding", "Self-serve", "Guided", "White-glove", "Custom"),
    ("SSO / SAML", "-", "-", "Y", "Y"),
    ("Custom Branding", "-", "-", "-", "Y"),
]

check_fill = PatternFill("solid", fgColor="DCFCE7")
no_fill = PatternFill("solid", fgColor="FEE2E2")
r = 4
for item in features:
    if item[0] is None:
        r += 1
        continue
    for j, val in enumerate(item, 1):
        if val is None:
            continue
        cell = ws5.cell(row=r, column=j, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if j > 1 else "left")
        if j > 1:
            if val == "Y":
                cell.fill = check_fill
                cell.value = "Yes"
            elif val == "-":
                cell.fill = no_fill
                cell.value = "-"
    if item[1] is None:
        ws5.cell(row=r, column=1).font = Font(bold=True, size=11, color="4F46E5")
        for jj in range(1, 6):
            ws5.cell(row=r, column=jj).fill = PatternFill("solid", fgColor="EEF2FF")
    r += 1
ws5.column_dimensions["A"].width = 30
for col_letter in ["B", "C", "D", "E"]:
    ws5.column_dimensions[col_letter].width = 18
# SHEET 6: Go-to-Market
ws6 = wb.create_sheet("Go-to-Market")
ws6.cell(row=1, column=1, value="GO-TO-MARKET STRATEGY").font = Font(bold=True, size=14, color="4F46E5")
gtm_headers = ["Phase", "Timeline", "Activities", "Budget", "KPIs"]
for j, h in enumerate(gtm_headers, 1):
    ws6.cell(row=3, column=j, value=h)
style_header(ws6, 3, 5)
gtm_data = [
    ["Phase 1: Validate", "Months 1-6", "Beta launch with 10-20 Italian factories\nContent marketing (lean blog, LinkedIn)\nLean consultant partnerships\nProduct-market fit validation", "$15,000", "20 beta users\n5 paying customers\nNPS > 40"],
    ["Phase 2: Scale EU", "Months 7-18", "Launch in IT, DE, UK, Balkans\nPaid ads (Google, LinkedIn)\nTrade show presence (SPS, Hannover)\nPartner channel development\nLocalize for DE, FR", "$120,000", "200 customers\n$50K MRR\nCAC < $500"],
    ["Phase 3: Expand", "Months 19-36", "Enter US market\nEnterprise sales team\nStrategic partnerships (ERP vendors)\nAdvanced AI features\nMobile app launch", "$500,000", "1,000 customers\n$300K MRR\nLTV/CAC > 3"],
    ["Phase 4: Dominate", "Months 37-60", "Market leadership in EU\nAcquisition targets\nIPO preparation\nPlatform ecosystem\nMarketplace for lean add-ons", "$1,500,000", "3,000+ customers\n$1M+ MRR\nProfitable"],
]
for i, row in enumerate(gtm_data, 4):
    for j, val in enumerate(row, 1):
        cell = ws6.cell(row=i, column=j, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
ws6.column_dimensions["A"].width = 22
ws6.column_dimensions["B"].width = 16
ws6.column_dimensions["C"].width = 45
ws6.column_dimensions["D"].width = 14
ws6.column_dimensions["E"].width = 22
for r_idx in range(4, 8):
    ws6.row_dimensions[r_idx].height = 80
# SHEET 7: KPIs & Milestones
ws7 = wb.create_sheet("KPIs & Milestones")
ws7.cell(row=1, column=1, value="KEY PERFORMANCE INDICATORS & MILESTONES").font = Font(bold=True, size=14, color="4F46E5")
for j, h in enumerate(years, 1):
    ws7.cell(row=3, column=j, value=h)
style_header(ws7, 3, 6)

kpis = [
    ("GROWTH METRICS", None),
    ("Active Customers", "ref"),
    ("MRR", "ref"),
    ("ARR", "ref"),
    ("YoY Revenue Growth", "calc"),
    (None, None),
    ("UNIT ECONOMICS", None),
    ("ARPU (Monthly)", [49, 95, 145, 185, 220]),
    ("CAC (Customer Acquisition Cost)", [800, 500, 400, 350, 300]),
    ("LTV (Lifetime Value)", "calc"),
    ("LTV/CAC Ratio", "calc"),
    ("Payback Period (Months)", "calc"),
    (None, None),
    ("OPERATIONAL", None),
    ("Monthly Churn Rate", [0.013, 0.010, 0.008, 0.007, 0.006]),
    ("NPS Score", [30, 45, 55, 60, 65]),
    ("Support Tickets/Customer/Mo", [2.5, 1.8, 1.2, 0.8, 0.5]),
    ("Uptime SLA", [0.995, 0.999, 0.999, 0.9999, 0.9999]),
]

r = 4
arpu_row = None
cac_row = None
ltv_row = None
arr_kpi_row = None
for item in kpis:
    label, values = item
    if label is None:
        r += 1
        continue
    ws7.cell(row=r, column=1, value=label).border = thin_border
    if label in ["GROWTH METRICS", "UNIT ECONOMICS", "OPERATIONAL"]:
        ws7.cell(row=r, column=1).font = Font(bold=True, size=11, color="4F46E5")
        ws7.cell(row=r, column=1).fill = PatternFill("solid", fgColor="EEF2FF")
    elif label == "Active Customers":
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws7.cell(row=r, column=col, value=f"='Revenue Projections'!{cc}{R_ACTIVE}").font = formula_font
            ws7.cell(row=r, column=col).number_format = num_fmt
            ws7.cell(row=r, column=col).border = thin_border
    elif label == "MRR":
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws7.cell(row=r, column=col, value=f"='Revenue Projections'!{cc}{R_TOTAL_MRR}").font = formula_font
            ws7.cell(row=r, column=col).number_format = money_fmt
            ws7.cell(row=r, column=col).border = thin_border
    elif label == "ARR":
        arr_kpi_row = r
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws7.cell(row=r, column=col, value=f"='Revenue Projections'!{cc}{R_ARR}").font = formula_font
            ws7.cell(row=r, column=col).number_format = money_fmt
            ws7.cell(row=r, column=col).border = thin_border
    elif label == "YoY Revenue Growth":
        for col in range(2, 7):
            cc = get_column_letter(col)
            if col == 2:
                ws7.cell(row=r, column=col, value="N/A").border = thin_border
            else:
                prev = get_column_letter(col - 1)
                ws7.cell(row=r, column=col, value=f"=IF({prev}{arr_kpi_row}=0,0,({cc}{arr_kpi_row}-{prev}{arr_kpi_row})/{prev}{arr_kpi_row})").font = formula_font
                ws7.cell(row=r, column=col).number_format = pct_fmt
                ws7.cell(row=r, column=col).border = thin_border
    elif label == "ARPU (Monthly)":
        arpu_row = r
        for j, v in enumerate(values, 2):
            cell = ws7.cell(row=r, column=j, value=v)
            cell.border = thin_border
            cell.font = input_font
            cell.number_format = money_fmt
    elif label == "CAC (Customer Acquisition Cost)":
        cac_row = r
        for j, v in enumerate(values, 2):
            cell = ws7.cell(row=r, column=j, value=v)
            cell.border = thin_border
            cell.font = input_font
            cell.number_format = money_fmt
    elif label == "LTV (Lifetime Value)":
        ltv_row = r
        churn_row = r + 4
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws7.cell(row=r, column=col, value=f"=IF({cc}{churn_row}=0,0,{cc}{arpu_row}/({cc}{churn_row}))").font = formula_font
            ws7.cell(row=r, column=col).number_format = money_fmt
            ws7.cell(row=r, column=col).border = thin_border
    elif label == "LTV/CAC Ratio":
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws7.cell(row=r, column=col, value=f"=IF({cc}{cac_row}=0,0,{cc}{ltv_row}/{cc}{cac_row})").font = formula_font
            ws7.cell(row=r, column=col).number_format = '0.0"x"'
            ws7.cell(row=r, column=col).border = thin_border
    elif label == "Payback Period (Months)":
        for col in range(2, 7):
            cc = get_column_letter(col)
            ws7.cell(row=r, column=col, value=f"=IF({cc}{arpu_row}=0,0,{cc}{cac_row}/{cc}{arpu_row})").font = formula_font
            ws7.cell(row=r, column=col).number_format = '0.0'
            ws7.cell(row=r, column=col).border = thin_border
    elif isinstance(values, list):
        for j, v in enumerate(values, 2):
            cell = ws7.cell(row=r, column=j, value=v)
            cell.border = thin_border
            cell.font = input_font
            if isinstance(v, float) and v < 1:
                cell.number_format = pct_fmt if v > 0.05 else '0.000%'
            elif isinstance(v, (int, float)) and v > 10:
                cell.number_format = num_fmt
            else:
                cell.number_format = '0.0'
    r += 1

# Milestones section
r += 2
ws7.cell(row=r, column=1, value="KEY MILESTONES").font = Font(bold=True, size=14, color="4F46E5")
r += 1
milestones = [
    ["Q2 2025", "Beta launch with 10 Italian pilot factories"],
    ["Q4 2025", "First 50 paying customers, break-even on infrastructure"],
    ["Q2 2026", "Launch in Germany and UK, 200 customers"],
    ["Q4 2026", "Reach $50K MRR, close seed round"],
    ["Q2 2027", "500 customers, mobile app launch"],
    ["Q4 2027", "Reach $300K MRR, Series A ready"],
    ["2028", "1,000+ customers, US expansion"],
    ["2029", "Market leader in EU SME lean SaaS, $1M+ MRR"],
    ["2030", "3,000+ customers, profitability, IPO/exit preparation"],
]
for ms in milestones:
    for j, val in enumerate(ms, 1):
        ws7.cell(row=r, column=j, value=val).border = thin_border
        if j == 1:
            ws7.cell(row=r, column=j).font = Font(bold=True, color="4F46E5")
    r += 1
auto_width(ws7)

# Save
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "LeanPilot_BusinessPlan.xlsx")
wb.save(output_path)
print(f"SUCCESS: Saved to {output_path}")