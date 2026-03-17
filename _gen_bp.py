"""Generate LeanPilot Business Plan Excel file with 7 formatted sheets."""
import sys, os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
TF = Font(name="Calibri", size=16, bold=True, color="1F4E79")
HF = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SF = Font(name="Calibri", size=12, bold=True, color="1F4E79")
DF = Font(name="Calibri", size=10)
BF = Font(name="Calibri", size=10, bold=True)
HFL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SFL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GFL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TB = Border(left=Side(style="thin",color="B4C6E7"),right=Side(style="thin",color="B4C6E7"),top=Side(style="thin",color="B4C6E7"),bottom=Side(style="thin",color="B4C6E7"))
CT = Alignment(horizontal="center", vertical="center")
WR = Alignment(wrap_text=True, vertical="top")

def hdr(ws, r, mc):
    for c in range(1, mc+1):
        cl = ws.cell(row=r, column=c); cl.font=HF; cl.fill=HFL; cl.alignment=CT; cl.border=TB

def sec(ws, r, mc):
    for c in range(1, mc+1):
        cl = ws.cell(row=r, column=c); cl.font=SF; cl.fill=SFL; cl.border=TB

def hl(ws, r, mc):
    for c in range(1, mc+1):
        cl = ws.cell(row=r, column=c); cl.font=BF; cl.fill=GFL

def wd(ws, data, sr=1):
    for ri, rd in enumerate(data, start=sr):
        for ci, v in enumerate(rd, start=1):
            cl = ws.cell(row=ri, column=ci, value=v); cl.font=DF; cl.border=TB; cl.alignment=WR

def aw(ws, mc):
    for c in range(1, mc+1):
        ml = 12
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value: ml = max(ml, min(len(str(cell.value)), 45))
        ws.column_dimensions[get_column_letter(c)].width = ml + 2

# === SHEET 1: Executive Summary ===
ws = wb.active; ws.title = "Executive Summary"
d = [
["LEANPILOT BUSINESS PLAN 2026-2030","","","",""],
["","","","",""],
["COMPANY OVERVIEW","","","",""],
["Company Name","LeanPilot SaaS Platform","","",""],
["Legal Entity","Centro Studi Grassi (Initial) -> LeanPilot S.r.l.","","",""],
["Founded","2025","","",""],
["Headquarters","Italy (EU)","","",""],
["Website","leanpilot.io","","",""],
["","","","",""],
["MISSION","","","",""],
["","Democratize Lean manufacturing for European SMEs with 17 integrated digital lean tools and an AI-powered factory copilot, replacing paper, spreadsheets, and expensive consultants.","","",""],
["","","","",""],
["PRODUCT SUMMARY","","","",""],
["Product Type","Cloud-based SaaS (B2B)","","",""],
["Core Platform","17 lean tools: OEE, 5 Why, Ishikawa, Kaizen, SMED, VSM, 6S, TPM, CILT, Pareto, A3, Andon, Gemba, Production/Downtime/Scrap, Hourly Tracking, Lean Assessment","","",""],
["AI Add-on (Pro)","Factory Copilot (GPT-4o), Root Cause AI, Auto Kaizen Suggestions","","",""],
["Tech Stack","Next.js 14 + FastAPI + PostgreSQL + OpenAI API + Docker","","",""],
["Languages","English, Italian (planned: German, French, Serbian)","","",""],
["GDPR","Fully compliant, EU data residency (Hetzner Cloud)","","",""],
["","","","",""],
["TARGET MARKET","","","",""],
["Primary","European SME manufacturers (10-500 employees)","","",""],
["Verticals","Automotive Tier 2/3, Food & Bev, Precision Engineering, Plastics, Metal, Pharma","","",""],
["Geography","Italy, Germany, France, Serbia/Balkans, UK, Austria, Poland","","",""],
["","","","",""],
["REVENUE MODEL","","","",""],
["Tier","Monthly Price","Target Segment","Key Value",""],
["Starter (Freemium)","Free","Lead gen, small workshops","Lean Assessment + basic tools, 1 user",""],
["Core","EUR 149/month","SMEs starting lean","All 17 lean tools, 10 users, dashboards",""],
["Pro","EUR 299/month","Advanced practitioners","Core + AI Copilot, Root Cause AI, unlimited users",""],
["Enterprise","Custom (EUR 599+/mo)","Large / multi-site","Pro + SSO, branding, dedicated support, API",""],
["","","","",""],
["KEY FINANCIAL TARGETS","","","",""],
["Metric","Year 1 (2026)","Year 3 (2028)","Year 5 (2030)",""],
["Total Active Customers","32","370","1,189",""],
["Annual Revenue (EUR)","37,520","662,144","2,366,144",""],
["MRR (EUR)","2,835","48,387","175,512",""],
["Gross Margin","78%","82%","85%",""],
["Net Margin","-10%","24.3%","49.2%",""],
["Break-even","Month 18-24 (mid-2027)","","",""],
["","","","",""],
["COMPETITIVE ADVANTAGES","","","",""],
["#","Advantage","","",""],
["1","Only affordable EU-focused Lean-complete SaaS platform","","",""],
["2","AI Copilot trained on real production data (not generic chatbot)","","",""],
["3","17 integrated lean tools vs competitors with 2-3 tools","","",""],
["4","Multi-language (EN/IT, expanding DE/FR/SR)","","",""],
["5","GDPR-compliant, EU data residency","","",""],
["6","5-minute setup vs months for enterprise solutions","","",""],
["7","10x-50x ROI: EUR 149/mo vs EUR 800-1500/day consultants","","",""],
]
wd(ws, d); ws.cell(1,1).font=TF
for r in [3,10,13,21,26,32,40]: sec(ws,r,5)
hdr(ws,27,5); hdr(ws,33,5); hdr(ws,41,5)
aw(ws,5); ws.column_dimensions["B"].width=55

# === SHEET 2: Market Analysis ===
ws = wb.create_sheet("Market Analysis")
d = [
["MARKET ANALYSIS - EU MANUFACTURING SMEs","","","",""],
["","","","",""],
["TOTAL ADDRESSABLE MARKET (TAM)","","","",""],
["Metric","Value","Source / Notes","",""],
["EU Manufacturing SMEs (10-500 emp)","~200,000","Eurostat 2024 NACE Rev.2 Section C","",""],
["Avg SaaS spend/factory/year","EUR 3,000","Mid-range Core+Pro pricing","",""],
["TAM (EUR)","EUR 600,000,000","200K x EUR 3,000","",""],
["","","","",""],
["SERVICEABLE ADDRESSABLE MARKET (SAM)","","","",""],
["Metric","Value","Notes","",""],
["Target countries (IT,DE,FR,UK,Balkans)","~85,000","42% of EU manufacturing SMEs","",""],
["Digitally ready factories","~35,000","~41% digital readiness (McKinsey)","",""],
["Avg annual spend","EUR 2,400","Weighted avg across tiers","",""],
["SAM (EUR)","EUR 84,000,000","35K x EUR 2,400","",""],
["","","","",""],
["SERVICEABLE OBTAINABLE MARKET (SOM)","","","",""],
["Metric","Value","Notes","",""],
["Year 5 customer base","3,500","~10% of SAM","",""],
["Blended ARPU/year","EUR 2,280","Core 65%, Pro 25%, Starter 10%","",""],
["SOM (EUR)","EUR 7,980,000","3,500 x EUR 2,280","",""],
["SOM % of SAM","9.5%","Achievable for category leader","",""],
["","","","",""],
["MARKET BY COUNTRY","","","",""],
["Country","Manuf. SMEs","Digital Ready %","Priority","Addressable"],
["Italy","45,000","35%","P1 - Home market","15,750"],
["Germany","35,000","52%","P1 - Largest EU mfg","18,200"],
["France","25,000","40%","P2 - Large market","10,000"],
["United Kingdom","20,000","48%","P2 - English-speaking","9,600"],
["Serbia/Balkans","8,000","28%","P2 - Cost advantage","2,240"],
["Poland","18,000","33%","P3 - Growth market","5,940"],
["Austria/Switzerland","6,000","50%","P3 - German premium","3,000"],
["TOTAL","157,000","~40% avg","","64,730"],
["","","","",""],
["INDUSTRY 4.0 TRENDS","","","",""],
["Metric","Value","Source","",""],
["Global I4.0 CAGR (2024-2030)","16.3%","MarketsandMarkets 2025","",""],
["EU Smart Manufacturing CAGR","14.2%","Grand View Research 2025","",""],
["SME digital transformation rate","38%","EU Commission 2025","",""],
["Manufacturing SaaS CAGR (EU)","18.5%","Statista 2025","",""],
["EU SMEs using cloud prod tools","24%","Eurostat ICT survey","",""],
["Expected cloud adoption by 2030","55%","EU Digital Decade targets","",""],
["AI in manufacturing","12% -> 45% by 2030","McKinsey Global Survey","",""],
["","","","",""],
["COMPETITIVE LANDSCAPE","","","",""],
["Competitor","Type","Price","Lean Tools","Weakness vs LeanPilot"],
["Siemens Opcenter","Enterprise MES","EUR 50K+/yr","Some","Too expensive for SMEs"],
["Retrocausal","AI Mfg (US)","USD 500+/mo","Limited","US-focused, no lean suite"],
["oee.ai","OEE SaaS (NL)","EUR 200+/mo","OEE only","Single tool, higher price"],
["Evocon","OEE SaaS (EE)","EUR 150+/mo","OEE+downtime","2-3 tools only"],
["Tulip.co","No-code MES (US)","Custom EUR 500+","Build-own","Heavy config, not lean-native"],
["MachineMetrics","IoT Analytics (US)","USD 500+/mo","OEE+analytics","Hardware dependent, US"],
["Paper/Spreadsheets","Manual","Free","Ad-hoc","No real-time, no AI, errors"],
["LEANPILOT","Lean SaaS+AI","EUR 0-299/mo","17 tools","Affordable, EU, comprehensive"],
]
wd(ws,d); ws.cell(1,1).font=TF
for r in [3,9,16,22,32,41]: sec(ws,r,5)
hdr(ws,4,5);hdr(ws,10,5);hdr(ws,17,5);hdr(ws,23,5);hdr(ws,33,5);hdr(ws,42,5)
aw(ws,5); ws.column_dimensions["C"].width=35; ws.column_dimensions["E"].width=35

# === SHEET 3: Revenue Projections ===
ws = wb.create_sheet("Revenue Projections")
d = [
["5-YEAR REVENUE PROJECTIONS","","","","","",""],
["","","","","","",""],
["CUSTOMER ACQUISITION","","","","","",""],
["Metric","Y1 (2026)","Y2 (2027)","Y3 (2028)","Y4 (2029)","Y5 (2030)","Notes"],
["New Starter (Free)",40,120,250,400,600,"Freemium funnel"],
["New Core",25,100,200,350,550,"Main revenue driver"],
["New Pro",10,40,100,200,350,"AI upsell from Core"],
["New Enterprise",0,5,15,30,50,"Direct sales Y2+"],
["Total New Customers",75,265,565,980,1550,""],
["","","","","","",""],
["CHURN ASSUMPTIONS","","","","","",""],
["Monthly Churn","7.0%","6.0%","5.5%","5.0%","4.5%","Improving"],
["Annual Retention","42%","48%","52%","54%","58%","=(1-churn)^12"],
["Enterprise Churn (annual)","3%","3%","3%","3%","3%","Contracted"],
["","","","","","",""],
["ACTIVE CUSTOMERS (after churn)","","","","","",""],
["Active Starter (Free)",17,65,157,277,451,""],
["Active Core",11,56,136,254,427,""],
["Active Pro",4,22,60,130,248,""],
["Active Enterprise",0,5,17,36,63,"3% annual churn"],
["TOTAL ACTIVE",32,148,370,697,1189,""],
["","","","","","",""],
["MONTHLY RECURRING REVENUE","","","","","",""],
["Starter MRR",0,0,0,0,0,"Free"],
["Core MRR (EUR 149/mo)",1639,8344,20264,37846,63623,""],
["Pro MRR (EUR 299/mo)",1196,6578,17940,38870,74152,""],
["Enterprise MRR (EUR 599/mo)",0,2995,10183,21564,37737,""],
["TOTAL MRR (EUR)",2835,17917,48387,98280,175512,""],
["","","","","","",""],
["ANNUAL RECURRING REVENUE","","","","","",""],
["ARR (EUR)",34020,215004,580644,1179360,2106144,"MRR x 12"],
["YoY Growth","-","532%","170%","103%","79%",""],
["","","","","","",""],
["ADDITIONAL REVENUE","","","","","",""],
["Onboarding Fees",3500,14500,31500,58000,90000,"Enterprise setup"],
["Consulting",0,10000,35000,75000,120000,"Implementation"],
["Annual Billing Premium",0,5000,15000,30000,50000,"Pre-payment float"],
["Total Additional",3500,29500,81500,163000,260000,""],
["","","","","","",""],
["TOTAL REVENUE","","","","","",""],
["Subscription (ARR)",34020,215004,580644,1179360,2106144,""],
["Additional Revenue",3500,29500,81500,163000,260000,""],
["TOTAL REVENUE (EUR)",37520,244504,662144,1342360,2366144,""],
["","","","","","",""],
["REVENUE MIX (% subscription)","","","","","",""],
["Core %","58%","47%","42%","39%","36%",""],
["Pro %","42%","37%","37%","40%","42%","Growing AI"],
["Enterprise %","0%","17%","21%","22%","22%",""],
]
wd(ws,d); ws.cell(1,1).font=TF
for r in [3,11,16,22,28,32,37,42]: sec(ws,r,7)
hdr(ws,4,7)
for r in [9,21,27,30,35,40]: hl(ws,r,7)
aw(ws,7)

# === SHEET 4: Cost Structure ===
ws = wb.create_sheet("Cost Structure")
d = [
["COST STRUCTURE & PROFITABILITY","","","","","",""],
["","","","","","",""],
["INFRASTRUCTURE (Monthly Avg)","","","","","",""],
["Item","Y1","Y2","Y3","Y4","Y5","Notes"],
["Cloud Hosting (Hetzner/AWS)",300,800,2000,4500,8000,"Scales w/ customers"],
["PostgreSQL DB",50,150,400,800,1500,"Managed DB"],
["CDN & Storage",30,80,200,400,700,"Assets"],
["Monitoring",50,100,200,300,400,"Sentry, Grafana"],
["Domain/SSL/Email",10,10,15,15,20,""],
["Monthly Total",440,1140,2815,6015,10620,""],
["Annual Total",5280,13680,33780,72180,127440,""],
["","","","","","",""],
["AI / API COSTS","","","","","",""],
["OpenAI API (monthly)",200,1500,5000,12000,25000,"Pro+Enterprise"],
["Annual AI Cost",2400,18000,60000,144000,300000,""],
["","","","","","",""],
["PAYMENT PROCESSING","","","","","",""],
["Stripe Fees (annual)",1088,6946,18767,38058,67168,"~2.9% revenue"],
["","","","","","",""],
["TEAM (Annual)","","","","","",""],
["Role","Y1","Y2","Y3","Y4","Y5","Hire"],
["CTO/Founder",0,40000,60000,75000,85000,"Salary Y2+"],
["Developer #1",0,45000,45000,50000,50000,"Mid-Y2"],
["Developer #2",0,0,45000,50000,50000,"Y3"],
["DevOps",0,0,0,45000,50000,"Y4"],
["Sales Manager",0,35000,45000,50000,55000,"Y2"],
["Sales Rep #2",0,0,35000,40000,45000,"Y3"],
["Customer Success",0,0,30000,35000,40000,"Y3"],
["Marketing Manager",0,0,0,40000,45000,"Y4"],
["Support Agent",0,0,0,25000,30000,"Y4"],
["Salaried Total",0,120000,260000,410000,450000,""],
["Contractors",15000,20000,15000,10000,10000,""],
["TOTAL PEOPLE",15000,140000,275000,420000,460000,""],
["","","","","","",""],
["MARKETING (Annual)","","","","","",""],
["Item","Y1","Y2","Y3","Y4","Y5",""],
["Content/SEO",2000,8000,15000,25000,35000,""],
["Paid Ads (Google/LinkedIn)",3000,15000,30000,50000,70000,""],
["Trade Shows",2000,8000,15000,25000,35000,""],
["Sales Tools",500,2000,3000,4000,5000,""],
["PR/Partnerships",0,3000,8000,12000,15000,""],
["Referral Program",0,2000,5000,10000,15000,""],
["TOTAL MARKETING",7500,38000,76000,126000,175000,""],
["","","","","","",""],
["OPERATIONAL (Annual)","","","","","",""],
["Legal/GDPR",3000,5000,8000,10000,12000,""],
["Accounting",2000,4000,6000,8000,10000,""],
["Insurance",1000,2000,3000,4000,5000,""],
["Office/Coworking",0,3000,6000,12000,18000,""],
["Travel",2000,5000,10000,15000,20000,""],
["Miscellaneous",2000,3000,5000,7000,8000,""],
["TOTAL OPERATIONAL",10000,22000,38000,56000,73000,""],
["","","","","","",""],
["=== COST SUMMARY (Annual) ===","","","","","",""],
["Category","Y1","Y2","Y3","Y4","Y5",""],
["Infrastructure",5280,13680,33780,72180,127440,""],
["AI/API",2400,18000,60000,144000,300000,""],
["Stripe",1088,6946,18767,38058,67168,""],
["People",15000,140000,275000,420000,460000,""],
["Marketing",7500,38000,76000,126000,175000,""],
["Operational",10000,22000,38000,56000,73000,""],
["TOTAL COSTS",41268,238626,501547,856238,1202608,""],
["","","","","","",""],
["=== PROFITABILITY ===","","","","","",""],
["","Y1","Y2","Y3","Y4","Y5",""],
["Revenue",37520,244504,662144,1342360,2366144,""],
["Costs",41268,238626,501547,856238,1202608,""],
["NET PROFIT/LOSS",-3748,5878,160597,486122,1163536,""],
["Net Margin","-10%","2.4%","24.3%","36.2%","49.2%",""],
["Cumulative P&L",-3748,2130,162727,648849,1812385,""],
["Burn Rate (monthly)",3439,19886,41796,71353,100217,""],
]
wd(ws,d); ws.cell(1,1).font=TF
for r in [3,13,16,19,33,40,47,53,59]: sec(ws,r,7)
hdr(ws,4,7);hdr(ws,20,7);hdr(ws,34,7);hdr(ws,41,7);hdr(ws,54,7);hdr(ws,60,7)
for r in [10,11,15,17,30,31,32,39,46,52,65,69,70]: hl(ws,r,7)
aw(ws,7)

# === SHEET 5: Pricing Strategy ===
ws = wb.create_sheet("Pricing Strategy")
d = [
["PRICING STRATEGY","","","","",""],
["","","","","",""],
["PRICING TIERS","","","","",""],
["Tier","EUR/month","Annual","Users","Target","Value"],
["Starter","Free","Free","1","Lead gen","Assessment + basic tools"],
["Core","149","1,788 (127 annual)","10","SMEs 10-100","All 17 lean tools"],
["Pro","299","3,588 (254 annual)","Unlimited","50-300 emp","Core + AI suite"],
["Enterprise","599+ custom","7,188+","Unlimited","Large/multi-site","Pro + SSO, API, dedicated"],
["","","","","",""],
["FEATURE MATRIX","","","","",""],
["Feature","Starter","Core","Pro","Enterprise",""],
["Lean Assessment","YES","YES","YES","YES",""],
["5 Why Analysis","3/month","Unlimited","Unlimited","Unlimited",""],
["Ishikawa","3/month","Unlimited","Unlimited","Unlimited",""],
["Kaizen Board","View only","Full","Full+AI","Custom",""],
["OEE Dashboard","NO","YES","YES","YES",""],
["Hourly Tracking","NO","YES","YES","YES",""],
["Andon Board","NO","YES","YES","YES",""],
["Pareto/A3","NO","YES","YES","YES",""],
["SMED Tracker","NO","YES","YES","YES",""],
["VSM","NO","YES","YES","YES",""],
["Gemba Walk","NO","YES","YES","YES",""],
["6S/TPM/CILT","NO","YES","YES","YES",""],
["AI Copilot","NO","NO","YES","YES",""],
["AI Root Cause","NO","NO","YES","YES",""],
["AI Auto Kaizen","NO","NO","YES","YES",""],
["Analytics","NO","NO","YES","YES",""],
["Custom Branding","NO","NO","YES","YES",""],
["SSO/SAML","NO","NO","NO","YES",""],
["API Access","NO","NO","NO","YES",""],
["Multi-site","NO","NO","NO","YES",""],
["Support","Community","Email 48h","Priority 24h","Dedicated AM",""],
["Languages","EN","EN+IT","EN+IT+DE+FR","All",""],
["Data Retention","30 days","1 year","3 years","Unlimited",""],
["","","","","",""],
["PRICE SENSITIVITY (EU SMEs)","","","","",""],
["Factor","Finding","Implication","","",""],
["SME SaaS budget","EUR 200-500/mo","Core EUR 149 is comfortable entry","","",""],
["Lean consultant rate","EUR 800-1,500/day","Pro annual = 2-4 consultant days","","",""],
["OEE improvement value","EUR 30K-100K/yr","15x-55x ROI on Core","","",""],
["IT/Balkans sensitivity","Higher","Freemium critical","","",""],
["DACH sensitivity","Lower, value-focused","Pro well-positioned","","",""],
["Monthly vs annual","70% prefer monthly","15% annual discount","","",""],
["AI premium","35% acceptable","EUR 150 uplift justified","","",""],
["","","","","",""],
["COMPETITIVE PRICING","","","","",""],
["Product","EUR/mo","Tools","AI","EU Focus","Assessment"],
["LeanPilot Core","149","17","No","YES","BEST VALUE lean SMEs"],
["LeanPilot Pro","299","17+AI","GPT-4o suite","YES","Unique AI+lean"],
["Evocon","~150+","2","No","YES","Same price, fewer tools"],
["oee.ai","~200+","1","Limited","YES","Higher, single tool"],
["Tulip.co","~500+","Build own","Limited","No","5x price, config heavy"],
["MachineMetrics","~500+ USD","OEE+","Some ML","No","Hardware, US"],
["Siemens","~4,000+","Full MES","Some","YES","20x cost, SME overkill"],
["","","","","",""],
["RECOMMENDATIONS","","","","",""],
["#","Action","Rationale","Impact","",""],
["1","Keep Starter free + 14-day trial","Reduces barrier in EU SME markets","Higher funnel","",""],
["2","Core at EUR 149/mo","Below EUR 200 threshold, competitive","Strong conversion","",""],
["3","Pro at EUR 299/mo","AI worth EUR 800+/mo consulting value","25%+ upsell","",""],
["4","Annual: Core 127, Pro 254/mo","15% discount improves retention","30-40% adoption","",""],
["5","Enterprise from EUR 599/mo","SSO, multi-site, API, support","Higher ARPU","",""],
["6","Per-user above 10 for Core","EUR 12/user/mo scales with factory","10-15% ARPU lift","",""],
]
wd(ws,d); ws.cell(1,1).font=TF
for r in [3,10,35,44,52,60]: sec(ws,r,6)
hdr(ws,4,6);hdr(ws,11,6);hdr(ws,36,6);hdr(ws,45,6);hdr(ws,53,6);hdr(ws,61,6)
aw(ws,6); ws.column_dimensions["C"].width=30; ws.column_dimensions["F"].width=30

# === SHEET 6: Go-to-Market ===
ws = wb.create_sheet("Go-to-Market")
d = [
["GO-TO-MARKET STRATEGY","","","",""],
["","","","",""],
["LAUNCH PHASES","","","",""],
["Phase","Timeline","Focus","Activities","Metrics"],
["Beta","Q1-Q2 2026","Validation","10-20 beta factories in Italy","NPS>40, 3 testimonials"],
["Italy Launch","Q3-Q4 2026","Home market","Public launch, MECSPE, consultants","50 paying, EUR 5K MRR"],
["DACH","Q1-Q2 2027","German mkts","DE language, SPS, VDMA","100 DACH customers"],
["Broader EU","Q3 2027-2028","FR/UK/Balkans","FR/SR language, partners","500+ total, EUR 50K MRR"],
["Scale","2029-2030","Leadership","Enterprise, IoT, Series A","2K+ cust, EUR 2M ARR"],
["","","","",""],
["CHANNELS","","","",""],
["Channel","Description","Cost","% Customers","Active From"],
["Inbound/SEO","Blog, SEO, free tools","Low-Med","40%","Q2 2026"],
["Outbound Sales","LinkedIn, email outreach","Medium","20%","Q1 2027"],
["Lean Consultants","50+ consultants refer clients","Low (15-20% rev share)","25%","Q3 2026"],
["Integrators","ERP/IT partners","Low-Med (20% share)","10%","Q3 2027"],
["Associations","CONFINDUSTRIA, VDMA","Low","5%","Q2 2026"],
["","","","",""],
["CONTENT MARKETING","","","",""],
["Type","Frequency","Purpose","Distribution",""],
["Blog: Lean Guides","2x/week","SEO + thought leadership","Website, LinkedIn, email",""],
["Case Studies","1x/month","Social proof","Website, sales, shows",""],
["OEE Calculator","Evergreen","Lead magnet","Website, ads",""],
["Lean Assessment","Evergreen","Lead qualification","Starter plan",""],
["YouTube","2x/month","Brand awareness","YouTube, LinkedIn",""],
["Webinars","1x/quarter","Lead gen","Zoom, LinkedIn",""],
["Whitepapers","1x/quarter","Enterprise leads","Gated download",""],
["Newsletter","1x/week","Nurture + retention","Brevo",""],
["","","","",""],
["SEO KEYWORDS","","","",""],
["Keyword","Volume/mo (EU)","Competition","Match",""],
["OEE software","3,200","Medium","Direct",""],
["lean manufacturing software","1,800","Medium","Core positioning",""],
["5S audit software","1,200","Low","6S audit",""],
["SMED software","800","Low","SMED tracker",""],
["kaizen board software","600","Low","Kaizen board",""],
["TPM software","900","Low-Med","TPM dashboard",""],
["root cause analysis tool","2,100","Medium","5Why+Ishikawa+AI",""],
["manufacturing AI","1,500","High","AI Copilot",""],
["continuous improvement sw","1,000","Medium","Full platform",""],
["value stream mapping tool","1,400","Medium","VSM feature",""],
["","","","",""],
["TRADE SHOWS","","","",""],
["Event","Location","Timing","Audience","Budget EUR"],
["MECSPE","Bologna, IT","March","Italian mfg","3K-5K"],
["SPS Nuremberg","Germany","November","Industry 4.0","5K-8K"],
["Hannover Messe","Germany","April","Global mfg","8K-15K"],
["Lean Summit DACH","Munich/Vienna","September","DACH lean","3K-5K"],
["Mfg Expo UK","Birmingham","June","UK mfg","3K-5K"],
["I4.0 Balkans","Belgrade","Various","Balkans mfg","1K-2K"],
["","","","",""],
["PARTNERSHIPS","","","",""],
["Type","Targets","Exchange","Revenue",""],
["Lean Consultants","50+ EU","Free Pro + certification","15-20% rev share",""],
["ERP Integrators","SAP B1, Odoo, Dynamics","Integration + co-mktg","20% rev share",""],
["Associations","CONFINDUSTRIA, VDMA","Content + events","Member discounts",""],
["Universities","PoliMi, TU Munich","Free academic licenses","Future pipeline",""],
["IoT Vendors","Sensor/PLC makers","Data integration","Bundled offers",""],
]
wd(ws,d); ws.cell(1,1).font=TF
for r in [3,11,18,28,39,49,55]: sec(ws,r,5)
hdr(ws,4,5);hdr(ws,12,5);hdr(ws,19,5);hdr(ws,29,5);hdr(ws,40,5);hdr(ws,50,5);hdr(ws,56,5)
aw(ws,5); ws.column_dimensions["D"].width=30

# === SHEET 7: KPIs & Milestones ===
ws = wb.create_sheet("KPIs & Milestones")
d = [
["KPIs, MILESTONES & RISKS","","","","",""],
["","","","","",""],
["KEY PERFORMANCE INDICATORS","","","","",""],
["KPI","Y1 Target","Y2 Target","Y3 Target","Measurement","Owner"],
["MRR (EUR)","2,835","17,917","48,387","Stripe","CEO"],
["New Signups/month","6","22","47","Analytics","Marketing"],
["Free-to-Paid Conv.","20%+","25%+","30%+","Cohort","Product"],
["Monthly Churn","<8%","<6%","<5.5%","Stripe","CS"],
["NRR",">90%",">100%",">110%","Revenue","Sales"],
["CAC (EUR)","<500","<400","<350","Spend/cust","Marketing"],
["LTV (EUR)",">1,500",">2,500",">4,000","ARPU/churn","Finance"],
["LTV:CAC",">3:1",">6:1",">10:1","Ratio","CEO"],
["DAU","30","200","600","PostHog","Product"],
["NPS",">30",">40",">50","Survey","CS"],
["Uptime","99.5%","99.9%","99.9%","Monitor","Eng"],
["AI Usage (Pro)","50 q/user","60 q/user","70 q/user","API logs","Product"],
["","","","","",""],
["PRODUCT MILESTONES","","","","",""],
["Quarter","Milestone","Description","Criteria","Priority",""],
["Q1 2026","Beta","10-20 Italian factories","10 users, <5% errors","P0",""],
["Q2 2026","Italy Launch","Public, Stripe billing","25 signups, 5 paying","P0",""],
["Q3 2026","AI Module","Copilot, RCA, Kaizen","10 Pro, >80% satisfaction","P0",""],
["Q4 2026","Analytics","Dashboards, PDF export","50+ customers","P1",""],
["Q1 2027","DACH Launch","German language","20 DACH signups","P1",""],
["Q2 2027","Mobile PWA","Shop floor mobile","30% mobile usage","P1",""],
["Q3 2027","Enterprise","SSO, API, multi-site","5 Enterprise deals","P1",""],
["Q4 2027","IoT","PLC/sensor auto OEE","10 connected factories","P2",""],
["Q1 2028","FR+SR","French/Serbian expansion","100+ non-IT/DE cust","P2",""],
["Q2 2028","Integrations","ERP, Zapier","3+ partners","P2",""],
["Q3 2028","Gemba Vision","AI quality inspection","5 pilots","P3",""],
["Q4 2028","Predictive","ML OEE prediction","20% Pro adoption","P3",""],
["","","","","",""],
["FUNDING","","","","",""],
["Stage","Amount EUR","Timeline","Use of Funds","Trigger","Type"],
["Bootstrap","20K-50K","2025-Q2 2026","MVP, hosting, legal","Founder savings","Founder"],
["Pre-Seed","100K-250K","Q3-Q4 2026","Hires, marketing","50 signups, PMF","Angels/grants"],
["Seed","500K-1M","Q2-Q3 2027","Team, EU growth","200 cust, 15K MRR","Seed VC"],
["Series A","2M-5M","2028-2029","Scale, IoT, 15+ team","1K cust, 100K MRR","Growth VC"],
["","","","","",""],
["TEAM PLAN","","","","",""],
["Dept","Y1","Y2","Y3","Y4","Y5"],
["Leadership",1,1,2,2,3],
["Engineering","0 (founder)",2,3,4,5],
["Sales",0,1,2,3,3],
["Marketing",0,0,0,1,2],
["CS/Support",0,0,1,2,3],
["Operations",0,0,0,0,1],
["TOTAL",1,4,8,12,17],
["Cost EUR","15,000","140,000","275,000","420,000","460,000"],
["","","","","",""],
["RISK REGISTER","","","","",""],
["Risk","Prob.","Impact","Mitigation","Owner",""],
["Slow SME adoption","Medium","High","Freemium, consultants, 14-day trial","CEO",""],
["High churn (>8%)","Med-High","High","Onboarding, CS, annual billing","Product",""],
["OpenAI cost increase","Low-Med","Medium","Multi-model, open-source, caps","CTO",""],
["Enterprise competition","Low","High","Speed, SME UX, lean-native, price","CEO",""],
["GDPR complexity","Medium","Medium","EU hosting, DPO, privacy-by-design","CTO",""],
["Funding gaps","Medium","High","Bootstrap path, revenue focus, grants","CEO",""],
["Key person risk","High","High","Document all, hire CTO early","CEO",""],
["Technical debt","Medium","Medium","Quarterly refactor, >70% test coverage","CTO",""],
["Market timing","Low-Med","Medium","Core works w/o AI, education marketing","CEO",""],
]
wd(ws,d); ws.cell(1,1).font=TF
for r in [3,18,33,38,43,51]: sec(ws,r,6)
hdr(ws,4,6);hdr(ws,19,6);hdr(ws,34,6);hdr(ws,39,6);hdr(ws,44,6);hdr(ws,52,6)
for r in [49,50]: hl(ws,r,6)
aw(ws,6); ws.column_dimensions["D"].width=40

# === SAVE ===
out = r"C:\Users\grass\Nextcloud4\Centro Studi Grassi\AI\Lean OS\LeanPilot_BusinessPlan_v2.xlsx"
wb.save(out)
print(f"OK: {out}")
print(f"Sheets: {wb.sheetnames}")
